#! python3
# -*- coding: utf-8 -*-
# author:Alexllll


 
import tkinter as tk  # 使用Tkinter前需要先导入
from tkinter import ttk
import openpyxl
import os
import pprint
from  tkinter.messagebox import showwarning , showinfo ,showerror,askyesno
import pickle
import sys
import time
import matplotlib.pyplot as plt
import matplotlib.backends.backend_tkagg
from random import randint





# 导入配置文件内容
# 定义变量 workbook的路径，名称 和 工作表
if os.path.exists('config.txt'):
    configFile = open('config.txt')
    configString = configFile.readlines()

    workbookPath = configString[0].split('::')[-1].strip()
    
    workbookSheet = configString[1].split('::')[-1].strip()

    workbookTakeout = configString[2].split('::')[-1].strip()

    configFile.close()

else:
    showerror(title = 'error' , message = 'no config file available !!!')

    sys.exit()

user = ''
 
# 第1步，实例化object，建立窗口window
window = tk.Tk()
 
# 第2步，给窗口的可视化起名字
window.title('备件管理')
 
# 第3步，设定窗口的大小(长 * 宽)
# w = window.winfo_screenwidth()
# h = window.winfo_screenheight()
# window.geometry("%dx%d" %(w,h))
#window.geometry('1000x600')  # 这里的乘是小x

window.wm_state( 'zoomed' )

# 更改icon
window.iconbitmap('gear.ico')

# 定义Login
""" def Login():
    
    addBtn.place(x = 20, y = 100 ,anchor = 'nw')
    delBtn.place(x = 20, y = 200 ,anchor = 'nw') """




# 定义Exit
def Exit():
    modifyBtn.place_forget()
    addBtn.place_forget()
    delBtn.place_forget()
    labelDelBtn.place_forget()
    labelModBtn.place_forget()
    settingMenu.entryconfig('Datebase',state = 'disabled')

    labelTakeFill.place_forget()
    e_takefill.place_forget()
    btnTakeout.place_forget()
    btnFillin.place_forget()

    global user
    user = ''
    


#创建一个frame

frame = tk.Frame(window)
frame.pack()

#frame 分成左右两半


 
# 第4步，在图形界面上创建一个标签用以显示内容并放置
l = tk.Label(frame , text='Line')
l2 = tk.Label(frame, text='Station')
l3 = tk.Label(frame, text='Part Description')
l4 = tk.Label(frame, text='Part No.')
l5 = tk.Label(frame, text='Company')
l6 = tk.Label(frame, text='Function')
l7 = tk.Label(frame, text='remark')
l8 = tk.Label(frame, text='Min Stock')
l9 = tk.Label(frame, text='Cur Stock')
l10 = tk.Label(frame, text='Cabinet')

l.grid(row =1 ,sticky = 'W')
l2.grid(row =2,sticky = 'W')
l3.grid(row =3 ,sticky = 'W')
l4.grid(row =4 ,sticky = 'W')
l5.grid(row =5 ,sticky = 'W')
l6.grid(row =6 ,sticky = 'W')
l7.grid(row =7 ,sticky = 'W')
l8.grid(row =8 ,sticky = 'W')
l9.grid(row =9 ,sticky = 'W')
l10.grid(row =10 ,sticky = 'W')
# 建立Entry 与上面的对应

#e = tk.Entry(frame, show = None)
e = ttk.Combobox(frame)
e['value'] = ('AG11','AG12','AG03','AH01','AH02','AL22','AL23','AL24','AL25','AL01','Noise Room')
#e.set('AG11')
var_e2 = tk.StringVar()
var_e3 = tk.StringVar()
var_e4 = tk.StringVar()
var_e5 = tk.StringVar()
var_e6 = tk.StringVar()
var_e7 = tk.StringVar()
var_e8 = tk.StringVar()
var_e9 = tk.StringVar()
var_e10 = tk.StringVar()
e2 = tk.Entry(frame,show = None,textvariable = var_e2)
e3 = tk.Entry(frame,show = None,textvariable = var_e3)
e4 = tk.Entry(frame,show = None,textvariable = var_e4)
e5 = tk.Entry(frame,show = None,textvariable = var_e5)
e6 = tk.Entry(frame,show = None,textvariable = var_e6)
e7 = tk.Entry(frame,show = None,textvariable = var_e7)
e8 = tk.Entry(frame,show = None,textvariable = var_e8)
e9 = tk.Entry(frame,show = None,textvariable = var_e9)
e10 = tk.Entry(frame,show = None,textvariable = var_e10)

e.grid(row =1 ,column = 1,sticky = 'W')
e2.grid(row =2 ,column = 1,sticky = 'W')
e3.grid(row =3 ,column = 1,sticky = 'W')
e4.grid(row =4 ,column = 1,sticky = 'W')
e5.grid(row =5 ,column = 1,sticky = 'W')
e6.grid(row =6 ,column = 1,sticky = 'W')
e7.grid(row =7 ,column = 1,sticky = 'W')
e8.grid(row =8 ,column = 1,sticky = 'W')
e9.grid(row =9 ,column = 1,sticky = 'W')
e10.grid(row =10 ,column = 1,sticky = 'W')

 

# 创建一个查询函数

def Query():
    criteria_1 = e.get()
    criteria_2 = e2.get()
    criteria_3 = e3.get()
    criteria_4 = e4.get()
    criteria_5 = e5.get()
    criteria_6 = e6.get()
    criteria_7 = e7.get()
    criteria_8 = e8.get()
    criteria_9 = e9.get()
    criteria_10 = e10.get()

    # 打开文件
    try :
        #os.chdir(workbookPath)
        workBook = openpyxl.load_workbook(workbookPath)
    except FileNotFoundError:
        showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
        return

    except openpyxl.utils.exceptions.InvalidFileException:
        showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
        return

        
    # 打开工作表
    try:
        worksheet = workBook[workbookSheet]
    except KeyError:
        showerror(title = '错误' , message = '工作表不存在')
        return

    lowStockNum = 0
    matches = []
    max_row = worksheet.max_row

    for rowNum in range(4, max_row + 1):
        if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
            highestRow = rowNum -1
            break
    
    # print(dimension)
    # print(max_row)
    for i in range(4, highestRow + 1):
        line = str(worksheet.cell(row = i, column = 1).value)
        station = str(worksheet.cell(row = i, column = 2).value)
        partDes = str(worksheet.cell(row = i, column = 3).value)
        partNo = str(worksheet.cell(row = i, column = 4).value)
        company = str(worksheet.cell(row = i, column = 5).value)
        function = str(worksheet.cell(row = i, column = 6).value)
        remark = str(worksheet.cell(row = i, column = 7).value)
        minStock = str(worksheet.cell(row = i, column = 8).value)
        curStock = str(worksheet.cell(row = i, column = 9).value)
        cabinet = str(worksheet.cell(row = i, column = 11).value)


        if int(curStock) < int(minStock) :
            lowStockNum += 1


        if criteria_1 !='' and (criteria_1.lower() not in line.lower()):
            pass
        elif criteria_2 != '' and (criteria_2.lower() not in station.lower()):
            pass
        elif criteria_3 !='' and (criteria_3.lower() not in partDes.lower()):
            pass
        elif criteria_4 !='' and (criteria_4.lower() not in partNo.lower()):
            pass
        elif criteria_5 !='' and (criteria_5.lower() not in company.lower()):
            pass
        elif criteria_6 !='' and (criteria_6.lower() not in function.lower()):
            pass
        elif criteria_7 !='' and (criteria_7.lower() not in remark.lower()):
            pass
        elif criteria_8 !='' and (criteria_8.lower() not in minStock.lower()):
            pass
        elif criteria_9 !='' and (criteria_9.lower() not in curStock.lower()):
            pass
        elif criteria_10 !='' and (criteria_10.lower() not in cabinet.lower()):
            pass
        elif line =='None' and station == 'None' and partDes == 'None' and partNo == 'None' and company == 'None' and function == 'None' and remark == 'None' and minStock == 'None' and curStock == 'None' and cabinet == 'None' :
            pass
        
        else :
            #matches.append({'Line': line ,'Station' : station ,'Part Description' : partDes , 'Part No.': partNo , 'Company' : company , 'Function': function , 'remark': remark, 'Min Stock': minStock, 'Cur Stock': curStock,'Cabinet': cabinet})
            matches.append([line , station , partDes ,  partNo ,  company ,  function , remark,  minStock, curStock, cabinet])

    # 关闭workBook
    #workBook.close()


    #pprint.pprint(matches)
    if lowStockNum > 0:
        showinfo(title = '库存不足' , message = '有' + str(lowStockNum) + '项库存不足')
    
    x = treeView.get_children()
    for item in x:
        treeView.delete(item)

    counter = 0
    for item in matches:
        counter += 1
        item.insert(0,counter)
        treeView.insert('', 'end' , text = 'line1'  , values = item)

def StockLowQuery():
    # 打开文件
    try :
        #os.chdir(workbookPath)
        workBook = openpyxl.load_workbook(workbookPath)
    except FileNotFoundError:
        showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
        return

    except openpyxl.utils.exceptions.InvalidFileException:
        showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
        return

        
    # 打开工作表
    try:
        worksheet = workBook[workbookSheet]
    except KeyError:
        showerror(title = '错误' , message = '工作表不存在')
        return

    matches =[]
    max_row = worksheet.max_row

    for rowNum in range(4, max_row + 1):
        if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
            highestRow = rowNum -1
            break
    
    
    for i in range(4, highestRow + 1):
        line = str(worksheet.cell(row = i, column = 1).value)
        station = str(worksheet.cell(row = i, column = 2).value)
        partDes = str(worksheet.cell(row = i, column = 3).value)
        partNo = str(worksheet.cell(row = i, column = 4).value)
        company = str(worksheet.cell(row = i, column = 5).value)
        function = str(worksheet.cell(row = i, column = 6).value)
        remark = str(worksheet.cell(row = i, column = 7).value)
        minStock = str(worksheet.cell(row = i, column = 8).value)
        curStock = str(worksheet.cell(row = i, column = 9).value)
        cabinet = str(worksheet.cell(row = i, column = 11).value)
        if int(curStock) < int(minStock) :
            matches.append([line , station , partDes ,  partNo ,  company ,  function , remark,  minStock, curStock, cabinet])

    x = treeView.get_children()
    for item in x:
        treeView.delete(item)

    counter = 0
    for item in matches:
        counter += 1
        item.insert(0,counter)
        treeView.insert('', 'end' , text = 'line1'  , values = item)


def Modify():
    line2mo = e.get()
    station2mo = e2.get()
    partDes2mo = e3.get()
    partNo2mo = e4.get()
    company2mo = e5.get()
    function2mo = e6.get()
    remark2mo = e7.get()
    minStock2mo = e8.get()
    curStock2mo = e9.get()
    cabinet2mo = e10.get()

    if line2mo =='' or station2mo =='' or partNo2mo =='':
        showerror(title = '错误' , message = 'line ,station and part NO 不能为空')
        return

    try:
    
        if int(minStock2mo) > 10 or int(minStock2mo) < 1:
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        elif int(curStock2mo) > 10 or int(curStock2mo) < 1:
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        elif int(curStock2mo) < int(minStock2mo):
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        else:
            pass


    except ValueError:

        showwarning(title = '警告',message = '库存请输入数字')
        return


    # 打开文件
    try :
        #os.chdir(workbookPath)
        workBook = openpyxl.load_workbook(workbookPath)
    except FileNotFoundError:
        showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
        return

    except openpyxl.utils.exceptions.InvalidFileException:
        showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
        return

        
    # 打开工作表
    try:
        worksheet = workBook[workbookSheet]
    except KeyError:
        showerror(title = '错误' , message = '工作表不存在')
        return


    matches = []
    max_row = worksheet.max_row

    for rowNum in range(4, max_row + 1):
        if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
            highestRow = rowNum -1
            break
    

    for i in range(4, highestRow + 1):
        partNo = str(worksheet.cell(row = i, column = 4).value)
        if partNo2mo == partNo:
            matches.append(i)
            
    if len(matches) == 0:
        showinfo(title = '提示' , message = '没有找到匹配条目')

    else: 

        for item in matches :
            worksheet.cell(row = item, column = 1).value = line2mo
            worksheet.cell(row = item, column = 2).value = station2mo
            worksheet.cell(row = item, column = 3).value = partDes2mo
            #worksheet.cell(row = highestRow + 1, column = 4).value = partNo2add
            worksheet.cell(row = item, column = 5).value = company2mo
            worksheet.cell(row = item, column = 6).value = function2mo
            worksheet.cell(row = item, column = 7).value = remark2mo
            worksheet.cell(row = item, column = 8).value = minStock2mo
            worksheet.cell(row = item, column = 9).value = curStock2mo
            worksheet.cell(row = item, column = 11).value = cabinet2mo


        try:
            workBook.save(workbookPath)
            showinfo(title = '添加成功' , message = '已成功修改' + str(len(matches)) + '项数据')

        except PermissionError:

            showerror(title = '添加失败' , message = 'Excel文件被打开，请先关闭，再重复此操作')
           
def Add():
    line2add = e.get()
    station2add = e2.get()
    partDes2add = e3.get()
    partNo2add = e4.get()
    company2add = e5.get()
    function2add = e6.get()
    remark2add = e7.get()
    minStock2add = e8.get()
    curStock2add = e9.get()
    cabinet2add = e10.get()

    if line2add =='' or station2add =='' or partNo2add =='':
        showerror(title = '错误' , message = 'line ,station and part NO 不能为空')
        return

    try:
    
        if int(minStock2add) > 10 or int(minStock2add) < 1:
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        elif int(curStock2add) > 10 or int(curStock2add) < 1:
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        elif int(curStock2add) < int(minStock2add):
            showwarning(title = '警告' , message = '1 <= 最小库存 <= 最大库存 <= 10')
            return

        else:
            pass


    except ValueError:

        showwarning(title = '警告',message = '库存请输入数字')
        return

    # 打开文件
    try :
        #os.chdir(workbookPath)
        workBook = openpyxl.load_workbook(workbookPath)
    except FileNotFoundError:
        showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
        return

    except openpyxl.utils.exceptions.InvalidFileException:
        showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
        return

    # 打开工作表
    try:
        worksheet = workBook[workbookSheet]
    except KeyError:
        showerror(title = '错误' , message = '工作表不存在')
        return


    
    max_row = worksheet.max_row

    for rowNum in range(4, max_row + 1):
        if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
            highestRow = rowNum -1
            break

    for i in range(4, highestRow + 1):
        # line = str(worksheet.cell(row = i, column = 1).value)
        # station = str(worksheet.cell(row = i, column = 2).value)
        partDes = str(worksheet.cell(row = i, column = 3).value)
        partNo = str(worksheet.cell(row = i, column = 4).value)
        # company = str(worksheet.cell(row = i, column = 5).value)
        # function = str(worksheet.cell(row = i, column = 6).value)
        # remark = str(worksheet.cell(row = i, column = 7).value)
        # minStock = str(worksheet.cell(row = i, column = 8).value)
        # curStock = str(worksheet.cell(row = i, column = 9).value)
        # cabinet = str(worksheet.cell(row = i, column = 11).value)

        if  partDes2add == partDes or partNo2add == partNo :
            showwarning(title = '警告' , message = '已存在，请重新输入(Part Description 和 Part No. 有重复)')
            break

        else:
            if i == highestRow:
                worksheet.cell(row = highestRow + 1, column = 1).value = line2add
                worksheet.cell(row = highestRow + 1, column = 2).value = station2add
                worksheet.cell(row = highestRow + 1, column = 3).value = partDes2add
                worksheet.cell(row = highestRow + 1, column = 4).value = partNo2add
                worksheet.cell(row = highestRow + 1, column = 5).value = company2add
                worksheet.cell(row = highestRow + 1, column = 6).value = function2add
                worksheet.cell(row = highestRow + 1, column = 7).value = remark2add
                worksheet.cell(row = highestRow + 1, column = 8).value = minStock2add
                worksheet.cell(row = highestRow + 1, column = 9).value = curStock2add
                worksheet.cell(row = highestRow + 1, column = 11).value = cabinet2add

                try:
                    workBook.save(workbookPath)
                    showinfo(title = '添加成功' , message = '已成功添加一项数据')

                except PermissionError:

                    showerror(title = '添加失败' , message = 'Excel文件被打开，请先关闭，再重复此操作')

                

            else:
                pass
            
def Del():
    # 根据partNo 来删除
    #delRows = 0
    partNo2del = e4.get()
    if partNo2del =='' :
        showinfo(title = '删除失败' , message = 'Part No. 不允许空白')
        return

    # 打开文件
    try :
        #os.chdir(workbookPath)
        workBook = openpyxl.load_workbook(workbookPath)
    except FileNotFoundError:
        showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
        return

    except openpyxl.utils.exceptions.InvalidFileException:
        showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
        return

    # 打开工作表
    try:
        worksheet = workBook[workbookSheet]
    except KeyError:
        showerror(title = '错误' , message = '工作表不存在')
        return
    
    max_row = worksheet.max_row
    rows_to_delete = []

    for rowNum in range(4, max_row + 1):
        if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
            highestRow = rowNum -1
            break
    
    for i in range(4, highestRow + 1):
        
        partNo = str(worksheet.cell(row = i, column = 4).value)
        

        if partNo2del == partNo :
            # worksheet.delete_rows(i)
            # delRows += 1
            rows_to_delete.append(i)

    lenRowsDel = len(rows_to_delete)

    if lenRowsDel == 0:
        showinfo(title = '删除失败' , message = '没有找到对应项')
    else :
        askDel = askyesno('删除确认？ ', '将删除'+str(lenRowsDel) + '项数据，是否继续？')
        if askDel == True:
            # for i in rows_to_delete:
            #     worksheet.delete_rows(i)
            # define a new list for the real delete rows ,cause the higher rows will collapse when delete the low rows .
            rowDel = []
            for i in range(lenRowsDel):
                rowDel.append(rows_to_delete[i] - i)

            for rownum in rowDel:
                worksheet.delete_rows(rownum)
        
            try:
                workBook.save(workbookPath)
                showinfo(title = '删除成功' , message = '成功删除'+str(lenRowsDel)+'项数据')
            except PermissionError:
                showerror(title = '添加失败' , message = 'Excel文件被打开，请先关闭，再重复此操作')

        else:
            pass

def setting():

    window_setting = tk.Toplevel(window)
    window_setting.geometry('500x300')
    window_setting.title('Datebase Setting')
    window_setting.iconbitmap('gear.ico')

    # 路径，工作表
    tk.Label(window_setting, text='文件名:').place(x=10, y=40)
    
    tk.Label(window_setting, text='备件表:').place(x=10, y=80)

    tk.Label(window_setting, text='领用记录表:').place(x=10, y=120)

    # 路径
    var_path = tk.StringVar()
    var_path.set(workbookPath)
    e_path = tk.Entry(window_setting, textvariable=var_path,width = 60)
    e_path.place(x = 120 , y = 40)

    

    # sheet
    var_sheet = tk.StringVar()
    var_sheet.set(workbookSheet)
    e_sheet = tk.Entry(window_setting, textvariable=var_sheet)
    e_sheet.place(x = 120 , y = 80)

     # take out sheet
    var_take = tk.StringVar()
    var_take.set(workbookTakeout)
    e_take = tk.Entry(window_setting, textvariable=var_take)
    e_take.place(x = 120 , y = 120)

    def config():
        path = e_path.get().strip()
        
        sheet = e_sheet.get().strip()

        takeout = e_take.get().strip()

        if path ==''  or sheet =='' or takeout =='':
            showerror(title = '错误' , message = '不允许空白')
            return

        else:
            setDic = {'path': path ,  'sheet' : sheet , 'take' : takeout}

            setFile = open('config.txt', 'w') 
            for k,v in setDic.items():
                setFile.write(k + ' :: ' + v +'\n')
            setFile.close()

            global workbookPath
            global workbookSheet
            global workbookTakeout

            workbookPath = path
            workbookSheet = sheet
            workbookTakeout = takeout

            window_setting.destroy()

     

    btnSet = tk.Button(window_setting , text = '确定' , command = config)
    btnSet.place(x = 200 , y = 210)


def user_login():
    # Add atop level for login
    window_login = tk.Toplevel(window)
    window_login.geometry('300x200')
    window_login.title('User Login')
    window_login.iconbitmap('gear.ico')
    window_login.wm_attributes('-topmost' , True)

    # 用户名和密码标签
    tk.Label(window_login, text='User name:').place(x=10, y=40)
    tk.Label(window_login, text='Password:').place(x=10, y=100)
    
    # 用户名
    var_usr_name = tk.StringVar()
    var_usr_name.set('')
    entry_usr_name = tk.Entry(window_login, textvariable=var_usr_name)
    entry_usr_name.place(x=120,y=40)
    # 用户密码
    var_usr_pwd = tk.StringVar()
    entry_usr_pwd = tk.Entry(window_login, textvariable=var_usr_pwd, show='*')
    entry_usr_pwd.place(x=120,y=100)

    def enterFun(self):
        btnLogin()

    def btnLogin():
        # 这两行代码就是获取用户输入的usr_name和usr_pwd
        usr_name = var_usr_name.get()
        usr_pwd = var_usr_pwd.get()

        # 这里设置异常捕获，当我们第一次访问用户信息文件时是不存在的，所以这里设置异常捕获。
        # 中间的两行就是我们的匹配，即程序将输入的信息和文件中的信息匹配。
        try:
            with open('usrs_info.pickle', 'rb') as usr_file:
                usrs_info = pickle.load(usr_file)
        except FileNotFoundError:
            # 这里就是我们在没有读取到`usr_file`的时候，程序会创建一个`usr_file`这个文件，并将管理员
            # 的用户和密码写入，即用户名为`admin`密码为`admin`。
            with open('usrs_info.pickle', 'wb') as usr_file:
                usrs_info = {'admin': 'admin'}
                pickle.dump(usrs_info, usr_file)
                usr_file.close()    # 必须先关闭，否则pickle.load()会出现EOFError: Ran out of input
        # 如果用户名和密码与文件中的匹配成功，则会登录成功，并跳出弹窗how are you? 加上你的用户名。
        if usr_name in usrs_info:
            if usr_pwd == usrs_info[usr_name]:
                # 弹出一个信息窗口
                # showinfo(title='Welcome', message='How are you? ' + usr_name)

                # 销毁窗口
                window_login.destroy()

                # 显示额外两个button
                addBtn.place(x = 20, y = 90 ,anchor = 'nw')
                modifyBtn.place(x = 20, y = 150 ,anchor = 'nw')
                labelModBtn.place(x = 105, y = 160)
                delBtn.place(x = 20, y = 210 ,anchor = 'nw')
                labelDelBtn.place(x = 105, y = 220)
                settingMenu.entryconfig('Datebase',state = 'active')

                labelTakeFill.place(x = 20 ,y = 5)
                e_takefill.place(x = 120 ,y = 5)
                btnTakeout.place(x = 20 ,y = 35 )
                btnFillin.place(x = 120  ,y = 35 )
                #print(addBtn.winfo_x())

                global user
                user = usr_name

            # 如果用户名匹配成功，而密码输入错误，则会弹出'Error, your password is wrong, try again.'
            else:
                showerror(message='Error, your password is wrong, try again.')
        else:  # 如果发现用户名不存在
            is_sign_up = askyesno('Welcome！ ', 'You have not sign up yet. Sign up now?')
            # 提示需不需要注册新用户
            if is_sign_up:
                btnSign()

            else:
                window_login.destroy()

        
    
    def btnSign():
        subSign = tk.Toplevel(window_login)
        subSign.geometry('300x200')
        subSign.title('Sign Up')
        subSign.iconbitmap('gear.ico')
        subSign.wm_attributes('-topmost' , True)
        

        new_name = tk.StringVar()  # 将输入的注册名赋值给变量
        new_name.set('')  # 将最初显示定为'example@python.com'
        tk.Label(subSign, text='User name: ').place(x=10, y=10)  # 将`User name:`放置在坐标（10,10）。
        entry_new_name = tk.Entry(subSign, textvariable=new_name)  # 创建一个注册名的`entry`，变量为`new_name`
        entry_new_name.place(x=130, y=10)  # `entry`放置在坐标（150,10）.

        new_pwd = tk.StringVar()
        tk.Label(subSign, text='Password: ').place(x=10, y=50)
        entry_usr_pwd = tk.Entry(subSign, textvariable=new_pwd, show='*')
        entry_usr_pwd.place(x=130, y=50)

        new_pwd_confirm = tk.StringVar()
        tk.Label(subSign, text='Confirm password: ').place(x=10, y=90)
        entry_usr_pwd_confirm = tk.Entry(subSign, textvariable=new_pwd_confirm, show='*')
        entry_usr_pwd_confirm.place(x=130, y=90)

        def sign():
             # 以下三行就是获取我们注册时所输入的信息
            np = new_pwd.get()
            npf = new_pwd_confirm.get()
            nn = new_name.get()
    
            # 这里是打开我们记录数据的文件，将注册信息读出
            with open('usrs_info.pickle', 'rb') as usr_file:
                exist_usr_info = pickle.load(usr_file)
            # 这里就是判断，如果两次密码输入不一致，则提示Error, Password and confirm password must be the same!
            if np != npf:
                showerror('Error', 'Password and confirm password must be the same!')
    
            # 如果用户名已经在我们的数据文件中，则提示Error, The user has already signed up!
            elif nn in exist_usr_info:
                showerror('Error', 'The user has already signed up!')
    
            # 最后如果输入无以上错误，则将注册输入的信息记录到文件当中，并提示注册成功Welcome！,You have successfully signed up!，然后销毁窗口。
            else:
                exist_usr_info[nn] = np
                with open('usrs_info.pickle', 'wb') as usr_file:
                    pickle.dump(exist_usr_info, usr_file)
                showinfo('Welcome', 'You have successfully signed up!')
                # 然后销毁窗口。
                subSign.destroy()



        # 下面的 sign_to_Hongwei_Website
        btn_comfirm_sign_up = tk.Button(subSign, text='Sign up', command=sign)
        btn_comfirm_sign_up.place(x=180, y=120)


    btn_Login = tk.Button(window_login, text='Login', command=btnLogin)
    btn_Sign= tk.Button(window_login, text='Sign Up', command=btnSign)

    btn_Login.place(x = 80, y = 150)
    btn_Sign.place(x = 140, y = 150)

    window_login.bind("<Return>",enterFun)


def Copy():
    sel = treeView.selection()
    
    if len(sel) != 1:
        pass
    else:
        item_text = treeView.item(sel[0],'values')
        e.set(item_text[1])
        var_e2.set(item_text[2])
        var_e3.set(item_text[3])
        var_e4.set(item_text[4])
        var_e5.set(item_text[5])
        var_e6.set(item_text[6])
        var_e7.set(item_text[7])
        var_e8.set(item_text[8])
        var_e9.set(item_text[9])
        var_e10.set(item_text[10])
        pass

def Clear():

    e.set('')
    var_e2.set('')
    var_e3.set('')
    var_e4.set('')
    var_e5.set('')
    var_e6.set('')
    var_e7.set('')
    var_e8.set('')
    var_e9.set('')
    var_e10.set('')

    x = treeView.get_children()
    for item in x:
        treeView.delete(item)


def Takeout():
    num = e_takefill.get().strip()
    partNo = e4.get().strip()
    try : 
        int(num)
    except ValueError:
        showerror(title = '错误' , message = '请输入数字')
        return
    if partNo =='' :
        showerror(title = '错误' , message = 'partNo 不能为空')
    else:
        confirm = askyesno('取料提醒 ', '领出' + num + '个part No 为' + partNo + '的备件')
        if confirm == True :
            # 打开文件
            try :
                #os.chdir(workbookPath)
                workBook = openpyxl.load_workbook(workbookPath)
            except FileNotFoundError:
                showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
                return

            except openpyxl.utils.exceptions.InvalidFileException:
                showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
                return

            # 打开工作表
            try:
                worksheet = workBook[workbookSheet]
                wbTakeout = workBook[workbookTakeout]
            except KeyError:
                showerror(title = '错误' , message = '工作表不存在')
                return
            
            max_row = worksheet.max_row
            max_row_take = wbTakeout.max_row
            
            # find the real highest row of  sheet
            for rowNum in range(4, max_row + 1):
                if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
                    highestRow = rowNum -1
                    break
            # find the real highest row of  sheet take out                        
            for rowNum in range(1, max_row_take + 1):
                if str(wbTakeout.cell(row = rowNum, column = 1).value) == 'None':
                    startrow = rowNum
                else :
                    startrow = max_row_take + 1
                    break
            
            take_num = 0
            for i in range(4, highestRow + 1):
                line = str(worksheet.cell(row = i, column = 1).value)
                partDes = str(worksheet.cell(row = i, column = 3).value)
                partNoDB = str(worksheet.cell(row = i, column = 4).value)
                curStock = str(worksheet.cell(row = i, column = 9).value)
                if partNo == partNoDB :
                    newStock = int(curStock) - int(num)
                    worksheet.cell(row = i, column = 9).value = newStock

                    wbTakeout.cell(row = startrow,column = 1 ).value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
                    wbTakeout.cell(row = startrow,column = 2 ).value = partDes
                    wbTakeout.cell(row = startrow,column = 3 ).value = partNo
                    wbTakeout.cell(row = startrow,column = 4 ).value = line
                    wbTakeout.cell(row = startrow,column = 5 ).value = int(num)
                    wbTakeout.cell(row = startrow,column = 6 ).value = user

                    take_num += 1
            
            if take_num == 0:
                showinfo(title = '领料失败' , message = '没有该料号')

            else :
            
                try:
                    workBook.save(workbookPath)
                    
                except PermissionError:
                    showerror(title = '领料失败' , message = 'Excel文件被打开，请先关闭，再重复此操作')
            


def Fillin():
    num = e_takefill.get().strip()
    partNo = e4.get().strip()
    try : 
        int(num)
    except ValueError:
        showerror(title = '错误' , message = '请输入数字')
        return
    if num =='' :
        pass
    else:
        confirm = askyesno('补料提醒 ', '存入' + num + '个part No 为' + partNo + '的备件')
        if confirm == True :
            # 打开文件
            try :
                #os.chdir(workbookPath)
                workBook = openpyxl.load_workbook(workbookPath)
            except FileNotFoundError:
                showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
                return

            except openpyxl.utils.exceptions.InvalidFileException:
                showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
                return

            # 打开工作表
            try:
                worksheet = workBook[workbookSheet]
                
            except KeyError:
                showerror(title = '错误' , message = '工作表不存在')
                return
            
            max_row = worksheet.max_row
            
            
            # find the real highest row of  sheet
            for rowNum in range(4, max_row + 1):
                if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
                    highestRow = rowNum -1
                    break
        
            
            for i in range(4, highestRow + 1):
                #line = str(worksheet.cell(row = i, column = 1).value)
                #partDes = str(worksheet.cell(row = i, column = 3).value)
                partNoDB = str(worksheet.cell(row = i, column = 4).value)
                curStock = str(worksheet.cell(row = i, column = 9).value)
                if partNo == partNoDB :
                    newStock = int(curStock) + int(num)
                    worksheet.cell(row = i, column = 9).value = newStock
            
            
            try:
                workBook.save(workbookPath)
                
            except PermissionError:
                showerror(title = '领料失败' , message = 'Excel文件被打开，请先关闭，再重复此操作')
    
def figure():
    window_figure = tk.Toplevel(window)
    window_figure.geometry('300x200')
    window_figure.title('图表显示')
    window_figure.iconbitmap('gear.ico')
    window_figure.wm_attributes('-topmost' ,True)

    label_title = tk.Label(window_figure , text='备件领用情况',font=('Arial', 20))
    label_title.pack(pady = 5 )

    frame_1 = tk.Frame(window_figure)
    frame_2 = tk.Frame(window_figure)

    frame_1.pack(pady = 20)
    frame_2.pack(pady = 20)

    e_fre = ttk.Combobox(frame_1,width = 10)
    e_fre['value'] = ('近一周','近一月','近二月','近三月','近半年','近一年')

    label = tk.Label(frame_1 , text='请选择期间')

    
    label.grid(row =1 ,sticky = 'W')
    e_fre.grid(row =1 ,column = 1,sticky = 'W')

    dic_offset = {'近一周' : 604800 ,'近一月' : 2592000 ,'近二月' : 5184000 ,'近三月' : 7776000 , '近半年' : 15552000 , '近一年' : 31104000}

    

    


    def show():
        plt.close()


        # 打开文件
        try :
            #os.chdir(workbookPath)
            workBook = openpyxl.load_workbook(workbookPath)
        except FileNotFoundError:
            showerror(title = '错误' , message = '找不到文件，请确认路径是否正确，文件是否存在')
            return

        except openpyxl.utils.exceptions.InvalidFileException:
            showerror(title = '错误' , message = '文件格式不对，需要xlsx扩展名')
            return

            
        # 打开工作表
        try:
            worksheet = workBook[workbookTakeout]
        except KeyError:
            showerror(title = '错误' , message = '工作表不存在')
            return

        

        name_list = []
        num_list = []
        
        dic_item = {}

        name_unique = []
        num_unique = []
        position_used = []
        max_row = worksheet.max_row

        # find the highest row
        
        for rowNum in range(2, max_row + 1):
            if str(worksheet.cell(row = rowNum, column = 1).value) == 'None':
                highestRow = rowNum -1

            else:
                highestRow = max_row


        # caculate the start time based on the selector
        # then can determin the start_row

        for k,v in dic_offset.items():
            if e_fre.get() == k:
                time_start = time.time()-v
                break

        try:

            for i in range(2,highestRow + 1):
                timef = str(worksheet.cell(row = i, column = 1).value)
                timefp = time.strptime(timef , '%Y-%m-%d %H:%M:%S')
                times = time.mktime(timefp)
                if times > time_start:
                    start_row = i
                    break
        except UnboundLocalError:
            showinfo(title = '提示' , message = '请选择期间')
            return

        #print(start_row)
        # prepare the plot data

        for i in range(start_row,highestRow + 1):
            name_list.append(str(worksheet.cell(row = i, column = 3).value))
            num_list.append(int(worksheet.cell(row = i, column = 5).value))

        for i in range(len(name_list)):
            if i in position_used:
                continue
            else:
                number = 0
                for j in range(len(name_list)):
                    if name_list[j] == name_list[i]:
                        position_used.append(j)
                        number += num_list[j]
                dic_item[name_list[i]] = number
        d_order = sorted(dic_item.items(),key = lambda x:x[1] , reverse = False)
        for x,y in d_order:
            name_unique.append(x)
            num_unique.append(y)

        # random color 

        def randomcolor():
            colorArr = ['1','2','3','4','5','6','7','8','9','A','B','C','D','E','F']
            color = ""
            for i in range(6):
                color += colorArr[randint(0,14)]
            return "#"+color  

        color = []

        for i in range(len(name_unique)):
            color.append(randomcolor())  

        p = plt.barh(name_unique, num_unique,height = 0.6 , align="center", color= color)
        

        for rect in p:
            w=rect.get_width()
            plt.text(w,rect.get_y()+rect.get_height()/2,'%d'%int(w),ha='left',va='center')
        plt.title('Spare Part Took out ')
        plt.xlabel("Number" )
        # plt.ylabel("part No" , verticalalignment = 'top' ,rotation = 'horizontal')
        plt.yticks(rotation = 45)

        plt.subplots_adjust(left=0.28, bottom=0.16 )

        plt.winter()

        plt.show()
        

        pass

    showBtn = tk.Button(frame_2, text = '显示' , width =10 ,height =2 , command = show)
    showBtn.pack(side = 'bottom' ,pady = 10)


# 创建一个menu
menubar = tk.Menu(window)

# 第6步，创建一个User菜单项（默认不下拉，下拉内容包括login 和 Exit功能项）
filemenu = tk.Menu(menubar, tearoff=0)
settingMenu = tk.Menu(menubar,tearoff = 0)
funcMenu = tk.Menu(menubar,tearoff = 0)
# 将上面定义的空菜单命名为File，放在菜单栏中，就是装入那个容器中
menubar.add_cascade(label='User', menu=filemenu)
menubar.add_cascade(label='Setting', menu=settingMenu)
menubar.add_cascade(label='Function', menu=funcMenu)

# 在File中加入login and  Exit 等小菜单，即我们平时看到的下拉菜单，每一个小菜单对应命令操作。
filemenu.add_command(label='Login', command=user_login)

filemenu.add_separator()    # 添加一条分隔线
filemenu.add_command(label='Exit', command=Exit) 

# 配置Setting菜单
settingMenu.add_command(label = 'Datebase' , command = setting , state='disabled')

# 配置Function菜单
funcMenu.add_command(label = 'figure' , command = figure )

# 创建菜单栏完成后，配置让菜单栏menubar显示出来
window.config(menu=menubar)   

# 新增一个frame
frame_m = tk.Frame(window)
frame_m.pack()

#新加一个button 用于查询
btnQuery = tk.Button(frame_m, text = '查询' , width =10 ,height =2 , command = Query)
btnQuery.pack(side = 'left' , padx = 5, pady = 5)

# 增加按钮
addBtn = tk.Button(window, text = '增加' , width =10 ,height =2 , command = Add)
# 修改按钮
modifyBtn = tk.Button(window, text = '修改' , width =10 ,height =2 , command = Modify)
labelModBtn = tk.Label(window, text = '（根据 Part No 修改)' )

# 删除按钮
delBtn = tk.Button(window, text = '删除' , width =10 ,height =2 , command = Del)
labelDelBtn = tk.Label(window, text = '（根据 Part No 删除)' )

# 拷贝按钮

btnCopy = tk.Button(frame_m, text = '拷贝' , width =10 ,height =2 , command = Copy)
btnCopy.pack(side = 'left', padx = 5, pady = 5)

# 清空按钮

btnClear = tk.Button(frame_m, text = '清空' , width =10 ,height =2 , command = Clear)
btnClear.pack(side = 'left', padx = 5, pady = 5)

# 库存不足显示

btnStockLow = tk.Button(frame_m, text = '库存不足查询' , width =15 ,height =2 , command = StockLowQuery)
btnStockLow.pack(side = 'left' ,padx = 10 ,pady =5 )



# label 领料/补料数量
labelTakeFill = tk.Label(window, text = '领料/补料数量 : ' )
#labelTakeFill.place(x = 20 ,y = 5)

e_takefill = tk.Entry(window, show =None ,width =10)
#e_takefill.place(x = 120 ,y = 5)

# 领料按钮
btnTakeout = tk.Button(window, text = '领料' ,bg = 'red' ,fg ='white', width =10 ,height =1 , command = Takeout)
#btnTakeout.place(x = 20 ,y = 35 )

# 补料按钮
btnFillin = tk.Button(window, text = '补料' ,bg = 'Green' ,fg ='white', width =10 ,height =1 , command = Fillin)
#btnFillin.place(x = 120  ,y = 35 )

 
# 创建查询结果显示框 
frame_2 = tk.Frame(window)
frame_2.pack(fill = 'both',expand = 1)
scrollBar = tk.Scrollbar(frame_2)
scrollBar.pack(side=tk.RIGHT, fill=tk.Y)


treeView = ttk.Treeview(frame_2, show = 'headings' , column = ('No.','Line' , 'Station' , 'Part Description' , 'Part NO.' , 'Company' , 'Function' , 'remark' , 'Min Stock' , 'Cur Stock' ,'Cabinet'))
treeView.column('No.',width = 40 , anchor = 'center')
treeView.column('Line',width = 100 , anchor = 'center')
treeView.column('Station',width = 100 , anchor = 'center')
treeView.column('Part Description',width = 200 , anchor = 'center')
treeView.column('Part NO.',width = 200 , anchor = 'center')
treeView.column('Company',width = 100 , anchor = 'center')
treeView.column('Function',width = 100 , anchor = 'center')
treeView.column('remark',width = 100 , anchor = 'center')
treeView.column('Min Stock',width = 50 , anchor = 'center')
treeView.column('Cur Stock',width = 50 , anchor = 'center')
treeView.column('Cabinet',width = 100 , anchor = 'center')

treeView.heading('No.' , text = 'No.')
treeView.heading('Line' , text = 'Line')
treeView.heading('Station' , text = 'Station')
treeView.heading('Part Description' , text = 'Part Description')
treeView.heading('Part NO.' , text = 'Part NO.')
treeView.heading('Company' , text = 'Company')
treeView.heading('Function' , text = 'Function')
treeView.heading('remark' , text = 'remark')
treeView.heading('Min Stock' , text = 'Min Stock')
treeView.heading('Cur Stock' , text = 'Cur Stock')
treeView.heading('Cabinet' , text = 'Cabinet')

treeView.pack(fill = 'both',expand = 1)


#Treeview组件与垂直滚动条结合
scrollBar.config(command=treeView.yview)

# 第12步，主窗口循环显示
window.mainloop()